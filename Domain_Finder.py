from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl import workbook, load_workbook

wb = openpyxl.load_workbook("Scraping.xlsx")
sheet1 = wb['Companies']

wb.create_sheet("Companies_domains")
sh = wb["Companies_domains"]
sh['A1'].value = "Company"
sh['B1'].value = "Domain"

row = sheet1.max_row
column = sheet1.max_column

title = []
link_loop= []

for i in range(2,row+1):
    for j in range(1, column + 1):
        company = sheet1.cell(i, j).value
        if company == None:
            continue
        #print(company)
        html_link = requests.get(f'https://www.google.com/search?q={company}').text
        soup = BeautifulSoup(html_link,'html.parser')
        #print(soup)
        with open('page.html','w') as html_file:
            html_file.write(html_link)
        titles = soup.find_all('div', {'class':'BNeawe vvjwJb AP7Wnd'})
        #print(titles)
        links = soup.findAll('div', {'class':'BNeawe UPmit AP7Wnd'})
        #print(links)
        all_title= []
        all_links = []
        for ttle in titles:
            tot_title = ttle.text
            all_title.append(tot_title)
        print(all_title)
        title.append(all_title[0])

        for lnks in links:
            tot_link = lnks.text
            all_links.append(tot_link)
        print(all_links)
        link = all_links[0]
        actual_link = link.split()[0]
        if 'www' in actual_link:
            link_loop.append(actual_link[4:])

        else:
            link_loop.append(actual_link)
print(title)
print(link_loop)

i = 2
for value in title:
    j = 1
    sh.cell(i, j).value = value
    i = i+1

i = 2
for value in link_loop:
    j = 2
    sh.cell(i, j).value = value
    i = i+1
wb.save("Scraping.xlsx")


